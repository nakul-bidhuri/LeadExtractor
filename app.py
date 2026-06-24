import streamlit as st
import sqlite3
import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
import time
import io
import json
from datetime import datetime, date
from urllib.parse import urljoin, urlparse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

# ─────────────────────── PAGE CONFIG ───────────────────────
st.set_page_config(
    page_title="LeadExtractor Pro",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────── CUSTOM CSS ───────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=Space+Grotesk:wght@400;500;600;700&display=swap');

/* Reset & Base */
* { box-sizing: border-box; }
html, body, .stApp { background: #06070d !important; }

.stApp {
    background: linear-gradient(135deg, #06070d 0%, #0d0f1a 50%, #06070d 100%) !important;
    font-family: 'Inter', sans-serif;
}

/* Hide default streamlit elements */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 1rem !important; max-width: 1400px; }

/* ── HERO HEADER ── */
.hero-header {
    background: linear-gradient(135deg, #0d0f1a 0%, #111827 100%);
    border: 1px solid rgba(99, 102, 241, 0.2);
    border-radius: 20px;
    padding: 2.5rem 3rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.hero-header::before {
    content: '';
    position: absolute;
    top: -50%;
    left: -10%;
    width: 60%;
    height: 200%;
    background: radial-gradient(ellipse, rgba(99,102,241,0.08) 0%, transparent 70%);
    pointer-events: none;
}
.hero-header::after {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, transparent, #6366f1, #8b5cf6, transparent);
}
.hero-title {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 2.4rem;
    font-weight: 700;
    color: #ffffff;
    margin: 0;
    letter-spacing: -0.02em;
}
.hero-title span { color: #6366f1; }
.hero-subtitle {
    color: #64748b;
    font-size: 0.95rem;
    margin-top: 0.4rem;
    font-weight: 400;
}
.hero-badge {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    background: rgba(99, 102, 241, 0.12);
    border: 1px solid rgba(99, 102, 241, 0.3);
    color: #818cf8;
    padding: 4px 12px;
    border-radius: 20px;
    font-size: 0.78rem;
    font-weight: 500;
    margin-top: 0.8rem;
}

/* ── STAT CARDS ── */
.stat-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 1rem;
    margin-bottom: 1.5rem;
}
.stat-card {
    background: #0d0f1a;
    border: 1px solid rgba(99,102,241,0.15);
    border-radius: 14px;
    padding: 1.2rem 1.4rem;
    position: relative;
    overflow: hidden;
    transition: border-color 0.2s ease, transform 0.2s ease;
}
.stat-card:hover {
    border-color: rgba(99,102,241,0.35);
    transform: translateY(-2px);
}
.stat-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(99,102,241,0.4), transparent);
}
.stat-icon {
    font-size: 1.4rem;
    margin-bottom: 0.5rem;
    display: block;
}
.stat-value {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 2rem;
    font-weight: 700;
    color: #fff;
    line-height: 1;
}
.stat-label {
    color: #64748b;
    font-size: 0.78rem;
    font-weight: 500;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-top: 0.3rem;
}
.stat-delta {
    color: #10b981;
    font-size: 0.75rem;
    margin-top: 0.3rem;
}

/* ── SECTION TITLES ── */
.section-title {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 1.1rem;
    font-weight: 600;
    color: #e2e8f0;
    margin: 0 0 1rem 0;
    display: flex;
    align-items: center;
    gap: 8px;
}
.section-title::after {
    content: '';
    flex: 1;
    height: 1px;
    background: rgba(99,102,241,0.2);
}

/* ── CARDS ── */
.card {
    background: #0d0f1a;
    border: 1px solid rgba(99,102,241,0.15);
    border-radius: 16px;
    padding: 1.5rem;
    margin-bottom: 1rem;
}

/* ── INPUT OVERRIDES ── */
.stTextInput input, .stTextArea textarea, .stSelectbox select {
    background: #111827 !important;
    border: 1px solid rgba(99,102,241,0.25) !important;
    border-radius: 10px !important;
    color: #e2e8f0 !important;
    font-family: 'Inter', sans-serif !important;
    transition: border-color 0.2s !important;
}
.stTextInput input:focus, .stTextArea textarea:focus {
    border-color: #6366f1 !important;
    box-shadow: 0 0 0 3px rgba(99,102,241,0.1) !important;
}

/* ── BUTTONS ── */
.stButton > button {
    background: linear-gradient(135deg, #6366f1, #8b5cf6) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.88rem !important;
    padding: 0.55rem 1.4rem !important;
    transition: all 0.2s ease !important;
    letter-spacing: 0.01em !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #4f46e5, #7c3aed) !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 15px rgba(99,102,241,0.35) !important;
}
.stButton > button:active { transform: translateY(0) !important; }

/* ── SIDEBAR ── */
[data-testid="stSidebar"] {
    background: #0a0b14 !important;
    border-right: 1px solid rgba(99,102,241,0.12) !important;
}
[data-testid="stSidebar"] .stMarkdown { color: #94a3b8 !important; }

.sidebar-logo {
    text-align: center;
    padding: 1.2rem 0 1rem;
    border-bottom: 1px solid rgba(99,102,241,0.15);
    margin-bottom: 1.2rem;
}
.sidebar-logo-text {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 1.3rem;
    font-weight: 700;
    color: #fff;
}
.sidebar-logo-text span { color: #6366f1; }
.sidebar-logo-sub {
    color: #475569;
    font-size: 0.72rem;
    margin-top: 2px;
    text-transform: uppercase;
    letter-spacing: 0.1em;
}

.nav-item {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 0.6rem 0.8rem;
    border-radius: 8px;
    color: #64748b;
    font-size: 0.88rem;
    font-weight: 500;
    cursor: pointer;
    transition: all 0.15s;
    margin-bottom: 2px;
}
.nav-item:hover { background: rgba(99,102,241,0.08); color: #94a3b8; }
.nav-item.active { background: rgba(99,102,241,0.15); color: #818cf8; border: 1px solid rgba(99,102,241,0.2); }

/* ── PROGRESS ── */
.progress-container {
    background: rgba(99,102,241,0.08);
    border: 1px solid rgba(99,102,241,0.2);
    border-radius: 12px;
    padding: 1.2rem 1.5rem;
    margin: 1rem 0;
}
.progress-text {
    color: #818cf8;
    font-size: 0.85rem;
    font-weight: 500;
    margin-bottom: 0.5rem;
}

/* ── TABLE ── */
.stDataFrame {
    border-radius: 12px !important;
    overflow: hidden;
    border: 1px solid rgba(99,102,241,0.15) !important;
}
.stDataFrame th {
    background: #0d0f1a !important;
    color: #818cf8 !important;
    font-size: 0.78rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.08em !important;
    font-weight: 600 !important;
}
.stDataFrame td {
    background: #06070d !important;
    color: #e2e8f0 !important;
    font-size: 0.85rem !important;
}

/* ── CHIPS ── */
.chip {
    display: inline-flex;
    align-items: center;
    gap: 4px;
    background: rgba(99,102,241,0.1);
    border: 1px solid rgba(99,102,241,0.2);
    color: #818cf8;
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 500;
}
.chip-green { background: rgba(16,185,129,0.1); border-color: rgba(16,185,129,0.2); color: #10b981; }
.chip-orange { background: rgba(245,158,11,0.1); border-color: rgba(245,158,11,0.2); color: #f59e0b; }

/* ── ALERT OVERRIDES ── */
.stSuccess { background: rgba(16,185,129,0.08) !important; border: 1px solid rgba(16,185,129,0.2) !important; color: #10b981 !important; border-radius: 10px !important; }
.stError { background: rgba(239,68,68,0.08) !important; border: 1px solid rgba(239,68,68,0.2) !important; color: #ef4444 !important; border-radius: 10px !important; }
.stWarning { background: rgba(245,158,11,0.08) !important; border: 1px solid rgba(245,158,11,0.2) !important; color: #f59e0b !important; border-radius: 10px !important; }
.stInfo { background: rgba(99,102,241,0.08) !important; border: 1px solid rgba(99,102,241,0.2) !important; color: #818cf8 !important; border-radius: 10px !important; }

/* ── TABS ── */
.stTabs [data-baseweb="tab-list"] {
    background: #0d0f1a !important;
    border-radius: 12px !important;
    padding: 4px !important;
    border: 1px solid rgba(99,102,241,0.15) !important;
    gap: 2px !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    color: #64748b !important;
    border-radius: 8px !important;
    font-size: 0.85rem !important;
    font-weight: 500 !important;
    padding: 6px 16px !important;
}
.stTabs [aria-selected="true"] {
    background: rgba(99,102,241,0.2) !important;
    color: #818cf8 !important;
}

/* ── LEAD CARD ── */
.lead-card {
    background: #0d0f1a;
    border: 1px solid rgba(99,102,241,0.12);
    border-radius: 12px;
    padding: 1rem 1.2rem;
    margin-bottom: 0.6rem;
    display: grid;
    grid-template-columns: 40px 1fr auto;
    gap: 1rem;
    align-items: center;
    transition: border-color 0.2s;
}
.lead-card:hover { border-color: rgba(99,102,241,0.3); }
.lead-avatar {
    width: 40px; height: 40px;
    background: linear-gradient(135deg, #6366f1, #8b5cf6);
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-weight: 700; color: white; font-size: 1rem;
}
.lead-domain { color: #e2e8f0; font-weight: 600; font-size: 0.9rem; }
.lead-email { color: #64748b; font-size: 0.8rem; margin-top: 2px; }
.lead-date { color: #475569; font-size: 0.75rem; text-align: right; }

/* ── PULSE ANIMATION ── */
@keyframes pulse-ring {
    0% { transform: scale(0.8); opacity: 1; }
    100% { transform: scale(1.4); opacity: 0; }
}
.pulse-dot {
    display: inline-block;
    width: 8px; height: 8px;
    background: #10b981;
    border-radius: 50%;
    position: relative;
}
.pulse-dot::before {
    content: '';
    position: absolute;
    top: 0; left: 0;
    width: 100%; height: 100%;
    border-radius: 50%;
    background: #10b981;
    animation: pulse-ring 1.5s ease-out infinite;
}

/* ── GRADIENT TEXT ── */
.gradient-text {
    background: linear-gradient(135deg, #6366f1, #8b5cf6, #a78bfa);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}

/* ── METRIC CARD ── */
[data-testid="metric-container"] {
    background: #0d0f1a !important;
    border: 1px solid rgba(99,102,241,0.15) !important;
    border-radius: 12px !important;
    padding: 1rem !important;
}
[data-testid="metric-container"] label { color: #64748b !important; }
[data-testid="metric-container"] [data-testid="stMetricValue"] { color: #e2e8f0 !important; }

/* ── DIVIDER ── */
hr { border-color: rgba(99,102,241,0.12) !important; }

/* ── SELECTBOX ── */
.stSelectbox [data-baseweb="select"] > div {
    background: #111827 !important;
    border-color: rgba(99,102,241,0.25) !important;
    color: #e2e8f0 !important;
    border-radius: 10px !important;
}

/* ── FILE UPLOADER ── */
[data-testid="stFileUploadDropzone"] {
    background: rgba(99,102,241,0.04) !important;
    border: 2px dashed rgba(99,102,241,0.25) !important;
    border-radius: 12px !important;
    color: #64748b !important;
}

/* ── CHECKBOX ── */
.stCheckbox label span { color: #94a3b8 !important; }

/* ── EXPANDER ── */
[data-testid="stExpander"] {
    background: #0d0f1a !important;
    border: 1px solid rgba(99,102,241,0.15) !important;
    border-radius: 12px !important;
}
[data-testid="stExpander"] summary { color: #e2e8f0 !important; }

/* ── SCROLLBAR ── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #06070d; }
::-webkit-scrollbar-thumb { background: rgba(99,102,241,0.3); border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: rgba(99,102,241,0.5); }
</style>
""", unsafe_allow_html=True)

# ─────────────────────── DATABASE ───────────────────────
DB_PATH = "leads.db"

def get_conn():
    return sqlite3.connect(DB_PATH, check_same_thread=False)

def init_db():
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            website TEXT,
            company TEXT,
            email TEXT,
            phone TEXT,
            linkedin TEXT,
            facebook TEXT,
            instagram TEXT,
            logo_url TEXT,
            ai_summary TEXT,
            date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # Migrate old DB if columns missing
    existing = [row[1] for row in c.execute("PRAGMA table_info(leads)").fetchall()]
    for col, dtype in [
        ("phone", "TEXT"), ("linkedin", "TEXT"), ("facebook", "TEXT"),
        ("instagram", "TEXT"), ("logo_url", "TEXT"), ("ai_summary", "TEXT")
    ]:
        if col not in existing:
            c.execute(f"ALTER TABLE leads ADD COLUMN {col} {dtype}")
    conn.commit()
    conn.close()

init_db()

# ─────────────────────── EXTRACTORS ───────────────────────
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

def normalize_url(url: str) -> str:
    url = url.strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    return url

def fetch_page(url: str, timeout: int = 12):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        r.raise_for_status()
        return r.text, r.url
    except Exception as e:
        return None, str(e)

def extract_emails(text: str) -> list:
    pattern = r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
    return list(set(re.findall(pattern, text)))

def extract_phones(text: str) -> list:
    pattern = r"(?:\+?\d[\d\s\-().]{7,}\d)"
    raw = re.findall(pattern, text)
    cleaned = []
    for p in raw:
        p = p.strip()
        digits = re.sub(r"\D", "", p)
        if 7 <= len(digits) <= 15:
            cleaned.append(p)
    return list(set(cleaned[:5]))

def extract_social(soup, base_url: str) -> dict:
    social = {"linkedin": "", "facebook": "", "instagram": ""}
    patterns = {
        "linkedin": r"linkedin\.com/(?:company|in)/[\w\-]+",
        "facebook": r"facebook\.com/[\w.]+",
        "instagram": r"instagram\.com/[\w.]+",
    }
    full_text = str(soup)
    for key, pat in patterns.items():
        m = re.search(pat, full_text)
        if m:
            social[key] = "https://www." + m.group(0)
    return social

def extract_logo(soup, base_url: str) -> str:
    for selector in [
        'link[rel="icon"]', 'link[rel="shortcut icon"]',
        'link[rel="apple-touch-icon"]', 'meta[property="og:image"]'
    ]:
        el = soup.select_one(selector)
        if el:
            href = el.get("href") or el.get("content", "")
            if href:
                return urljoin(base_url, href)
    # Try img with "logo" in class/alt/src
    for img in soup.find_all("img"):
        src = img.get("src", "")
        alt = img.get("alt", "").lower()
        cls = " ".join(img.get("class", [])).lower()
        if "logo" in src.lower() or "logo" in alt or "logo" in cls:
            return urljoin(base_url, src)
    return ""

def extract_company(soup, url: str) -> str:
    # og:site_name → title → domain
    og = soup.find("meta", property="og:site_name")
    if og and og.get("content"):
        return og["content"].strip()
    t = soup.find("title")
    if t and t.text:
        name = t.text.strip().split("|")[0].split("-")[0].strip()
        if len(name) < 60:
            return name
    return urlparse(url).netloc.replace("www.", "").split(".")[0].title()

def crawl_pages(base_url: str, max_pages: int = 5) -> list:
    """Crawl multiple pages of a site and return all HTML text."""
    visited = set()
    to_visit = [base_url]
    pages_text = []

    contact_keywords = ["contact", "about", "team", "support", "help", "reach"]

    while to_visit and len(visited) < max_pages:
        url = to_visit.pop(0)
        if url in visited:
            continue
        visited.add(url)
        html, _ = fetch_page(url)
        if not html:
            continue
        pages_text.append(html)
        if len(visited) >= max_pages:
            break
        # Find internal links with contact-like keywords
first_soup = BeautifulSoup(pages[0], "html.parser")
for a in soup.find_all("a", href=True):
            href = a["href"]
            full = urljoin(base_url, href)
            parsed = urlparse(full)
            if parsed.netloc == urlparse(base_url).netloc and full not in visited:
                text_lower = (a.get_text() + href).lower()
                if any(k in text_lower for k in contact_keywords):
                    to_visit.insert(0, full)
                elif len(to_visit) < 10:
                    to_visit.append(full)
    return pages_text

def extract_lead(url: str, multi_page: bool = True, max_pages: int = 3) -> dict:
    url = normalize_url(url)
    result = {
        "website": url, "company": "", "email": "", "phone": "",
        "linkedin": "", "facebook": "", "instagram": "",
        "logo_url": "", "ai_summary": "", "error": None
    }

    if multi_page:
        pages = crawl_pages(url, max_pages=max_pages)
    else:
        html, err = fetch_page(url)
        pages = [html] if html else []
        if not html:
            result["error"] = err
            return result

    if not pages:
        result["error"] = "Could not fetch page"
        return result

    all_text = " ".join(pages)
    first_soup = BeautifulSoup(pages[0], ""html.parser"")

    result["company"] = extract_company(first_soup, url)
    result["logo_url"] = extract_logo(first_soup, url)

    emails = extract_emails(all_text)
    # Filter obvious false positives
    emails = [e for e in emails if not any(x in e.lower() for x in ["@example", "@test", "@domain", "noreply@", "no-reply@", "@sentry", "@email.com"])]
    result["email"] = ", ".join(emails[:5]) if emails else ""

    phones = extract_phones(all_text)
    result["phone"] = ", ".join(phones[:3]) if phones else ""

    social = extract_social(first_soup, url)
    result.update(social)

    return result

# ─────────────────────── DATABASE HELPERS ───────────────────────
def save_lead(data: dict):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        INSERT INTO leads (website, company, email, phone, linkedin, facebook, instagram, logo_url, ai_summary)
        VALUES (?,?,?,?,?,?,?,?,?)
    """, (
        data.get("website",""), data.get("company",""), data.get("email",""),
        data.get("phone",""), data.get("linkedin",""), data.get("facebook",""),
        data.get("instagram",""), data.get("logo_url",""), data.get("ai_summary","")
    ))
    conn.commit()
    conn.close()

def load_leads(search="", date_from=None, date_to=None) -> pd.DataFrame:
    conn = get_conn()
    query = "SELECT * FROM leads WHERE 1=1"
    params = []
    if search:
        query += " AND (website LIKE ? OR company LIKE ? OR email LIKE ? OR phone LIKE ?)"
        s = f"%{search}%"
        params.extend([s, s, s, s])
    if date_from:
        query += " AND DATE(date) >= ?"
        params.append(str(date_from))
    if date_to:
        query += " AND DATE(date) <= ?"
        params.append(str(date_to))
    query += " ORDER BY date DESC"
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df

def delete_lead(lead_id: int):
    conn = get_conn()
    conn.execute("DELETE FROM leads WHERE id=?", (lead_id,))
    conn.commit()
    conn.close()

def update_lead(lead_id: int, company: str, email: str, phone: str):
    conn = get_conn()
    conn.execute(
        "UPDATE leads SET company=?, email=?, phone=? WHERE id=?",
        (company, email, phone, lead_id)
    )
    conn.commit()
    conn.close()

def get_stats() -> dict:
    conn = get_conn()
    c = conn.cursor()
    total = c.execute("SELECT COUNT(*) FROM leads").fetchone()[0]
    with_email = c.execute("SELECT COUNT(*) FROM leads WHERE email != ''").fetchone()[0]
    with_phone = c.execute("SELECT COUNT(*) FROM leads WHERE phone != ''").fetchone()[0]
    with_social = c.execute("SELECT COUNT(*) FROM leads WHERE linkedin != '' OR facebook != '' OR instagram != ''").fetchone()[0]
    today = c.execute("SELECT COUNT(*) FROM leads WHERE DATE(date)=DATE('now')").fetchone()[0]
    conn.close()
    return {"total": total, "with_email": with_email, "with_phone": with_phone, "with_social": with_social, "today": today}

def get_daily_counts() -> pd.DataFrame:
    conn = get_conn()
    df = pd.read_sql_query(
        "SELECT DATE(date) as day, COUNT(*) as count FROM leads GROUP BY DATE(date) ORDER BY day DESC LIMIT 30",
        conn
    )
    conn.close()
    return df

# ─────────────────────── EXPORT HELPERS ───────────────────────
def to_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_fill = PatternFill("solid", fgColor="1e1b4b")
    header_font = Font(bold=True, color="818cf8", size=10)
    border = Border(
        bottom=Side(style="thin", color="312e81"),
        right=Side(style="thin", color="312e81")
    )

    cols = list(df.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(1, ci, col.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    alt_fill = PatternFill("solid", fgColor="0f172a")
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, str(val) if val else "")
            if ri % 2 == 0:
                cell.fill = alt_fill
            cell.font = Font(color="cbd5e1", size=9)
            cell.border = border

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    stats = get_stats()
    ws2["A1"] = "LeadExtractor Pro — Summary"
    ws2["A1"].font = Font(bold=True, size=14, color="6366f1")
    ws2["A3"] = "Total Leads"; ws2["B3"] = stats["total"]
    ws2["A4"] = "With Email"; ws2["B4"] = stats["with_email"]
    ws2["A5"] = "With Phone"; ws2["B5"] = stats["with_phone"]
    ws2["A6"] = "With Social"; ws2["B6"] = stats["with_social"]
    ws2["A7"] = "Extracted Today"; ws2["B7"] = stats["today"]
    ws2["A9"] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws2["A9"].font = Font(color="64748b", size=9)

    wb.save(buf)
    return buf.getvalue()

def to_pdf_bytes(df: pd.DataFrame) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=15*mm, bottomMargin=15*mm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("title", fontSize=16, textColor=colors.HexColor("#6366f1"),
                                     fontName="Helvetica-Bold", alignment=TA_LEFT)
        sub_style = ParagraphStyle("sub", fontSize=9, textColor=colors.HexColor("#64748b"),
                                   fontName="Helvetica", alignment=TA_LEFT)

        story = [
            Paragraph("⚡ LeadExtractor Pro — Leads Export", title_style),
            Spacer(1, 4*mm),
            Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}  •  Total Records: {len(df)}", sub_style),
            Spacer(1, 6*mm),
        ]

        show_cols = ["id", "company", "website", "email", "phone", "linkedin", "date"]
        show_cols = [c for c in show_cols if c in df.columns]
        table_df = df[show_cols].fillna("").astype(str)

        data = [show_cols] + table_df.values.tolist()
        col_widths = [15*mm, 40*mm, 55*mm, 55*mm, 30*mm, 55*mm, 28*mm][:len(show_cols)]

        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1e1b4b")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#818cf8")),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,0), 8),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#06070d"), colors.HexColor("#0d0f1a")]),
            ("TEXTCOLOR", (0,1), (-1,-1), colors.HexColor("#cbd5e1")),
            ("FONTSIZE", (0,1), (-1,-1), 7),
            ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#1e1b4b")),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("ROWHEIGHT", (0,0), (-1,-1), 14),
            ("WORDWRAP", (0,0), (-1,-1), True),
        ]))
        story.append(t)
        doc.build(story)
        return buf.getvalue()
    except ImportError:
        return b""

# ─────────────────────── AI SUMMARY ───────────────────────
def get_ai_summary(company: str, website: str, emails: str) -> str:
    """Call Anthropic API for a company summary."""
    try:
        payload = {
            "model": "claude-sonnet-4-6",
            "max_tokens": 200,
            "messages": [{
                "role": "user",
                "content": (
                    f"Write a 2-3 sentence professional company summary for:\n"
                    f"Company: {company}\nWebsite: {website}\nContact: {emails}\n\n"
                    "Be concise. Include what they likely do and their industry."
                )
            }]
        }
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type": "application/json"},
            json=payload,
            timeout=20
        )
        data = r.json()
        if data.get("content"):
            return data["content"][0].get("text", "")
    except Exception:
        pass
    return ""

# ─────────────────────── SIDEBAR ───────────────────────
with st.sidebar:
    st.markdown("""
    <div class="sidebar-logo">
        <div class="sidebar-logo-text">Lead<span>Extractor</span></div>
        <div class="sidebar-logo-sub">Pro · v5.0</div>
    </div>
    """, unsafe_allow_html=True)

    page = st.radio(
        "Navigation",
        ["⚡ Extract", "📊 Dashboard", "🗂 Leads", "📅 Analytics", "⚙️ Settings"],
        label_visibility="collapsed"
    )

    st.divider()
    stats = get_stats()
    st.markdown(f"""
    <div style="padding: 0 0.5rem;">
        <div style="color:#475569; font-size:0.72rem; text-transform:uppercase; letter-spacing:0.08em; margin-bottom:0.6rem;">Quick Stats</div>
        <div style="display:flex; justify-content:space-between; margin-bottom:4px;">
            <span style="color:#64748b; font-size:0.8rem;">Total Leads</span>
            <span style="color:#818cf8; font-weight:600; font-size:0.8rem;">{stats['total']}</span>
        </div>
        <div style="display:flex; justify-content:space-between; margin-bottom:4px;">
            <span style="color:#64748b; font-size:0.8rem;">With Email</span>
            <span style="color:#10b981; font-weight:600; font-size:0.8rem;">{stats['with_email']}</span>
        </div>
        <div style="display:flex; justify-content:space-between; margin-bottom:4px;">
            <span style="color:#64748b; font-size:0.8rem;">With Phone</span>
            <span style="color:#f59e0b; font-weight:600; font-size:0.8rem;">{stats['with_phone']}</span>
        </div>
        <div style="display:flex; justify-content:space-between;">
            <span style="color:#64748b; font-size:0.8rem;">Today</span>
            <span style="color:#e2e8f0; font-weight:600; font-size:0.8rem;">{stats['today']}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown('<div style="color:#475569; font-size:0.7rem; text-align:center;">Built by Nakul Bidhuri 🚀</div>', unsafe_allow_html=True)

# ─────────────────────── HERO ───────────────────────
st.markdown("""
<div class="hero-header">
    <div class="hero-title">⚡ Lead<span>Extractor</span> Pro</div>
    <div class="hero-subtitle">AI-powered lead intelligence — emails, phones, social profiles, and more</div>
    <div class="hero-badge"><span class="pulse-dot"></span> Live · v5.0</div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────── PAGES ───────────────────────

# ══════════════════════════════════════════════
#  EXTRACT PAGE
# ══════════════════════════════════════════════
if page == "⚡ Extract":
    tab1, tab2 = st.tabs(["  Single URL  ", "  Bulk CSV Upload  "])

    with tab1:
        st.markdown('<p class="section-title">🎯 Extract from URL</p>', unsafe_allow_html=True)
        col1, col2 = st.columns([3, 1])
        with col1:
            url_input = st.text_input("Website URL", placeholder="e.g. https://example.com or example.com", label_visibility="collapsed")
        with col2:
            multi_page = st.checkbox("Multi-page crawl", value=True, help="Crawl contact/about pages too")
        
        col_a, col_b, col_c = st.columns([1, 1, 2])
        with col_a:
            ai_summary = st.checkbox("🤖 AI Summary", value=False, help="Generate company summary with AI")
        with col_b:
            max_p = st.selectbox("Max pages", [2, 3, 5, 8], index=1)

        if st.button("⚡ Extract Lead", use_container_width=True):
            if not url_input.strip():
                st.error("Please enter a URL.")
            else:
                with st.spinner(""):
                    prog = st.progress(0)
                    status = st.empty()

                    status.markdown('<div class="progress-container"><div class="progress-text">🌐 Fetching website...</div></div>', unsafe_allow_html=True)
                    prog.progress(20)
                    time.sleep(0.3)

                    status.markdown('<div class="progress-container"><div class="progress-text">🔍 Extracting data...</div></div>', unsafe_allow_html=True)
                    prog.progress(50)

                    result = extract_lead(url_input, multi_page=multi_page, max_pages=max_p)

                    prog.progress(75)
                    status.markdown('<div class="progress-container"><div class="progress-text">🤖 AI analysis...</div></div>', unsafe_allow_html=True)

                    if ai_summary and result.get("company") and not result.get("error"):
                        result["ai_summary"] = get_ai_summary(result["company"], result["website"], result["email"])

                    prog.progress(100)
                    status.empty()
                    prog.empty()

                if result.get("error"):
                    st.error(f"❌ {result['error']}")
                else:
                    save_lead(result)
                    st.success("✅ Lead extracted and saved!")

                    # Result card
                    c1, c2, c3 = st.columns(3)
                    with c1:
                        st.metric("Company", result["company"] or "—")
                    with c2:
                        st.metric("Emails Found", len([e for e in result["email"].split(",") if e.strip()]) if result["email"] else 0)
                    with c3:
                        st.metric("Phones Found", len([p for p in result["phone"].split(",") if p.strip()]) if result["phone"] else 0)

                    st.markdown('<div class="card">', unsafe_allow_html=True)
                    if result["logo_url"]:
                        st.markdown(f'<img src="{result["logo_url"]}" style="height:40px; border-radius:8px; margin-bottom:8px;" onerror="this.style.display=\'none\'">', unsafe_allow_html=True)

                    fields = {
                        "🌐 Website": result["website"],
                        "🏢 Company": result["company"],
                        "📧 Email(s)": result["email"] or "—",
                        "📱 Phone(s)": result["phone"] or "—",
                        "💼 LinkedIn": result["linkedin"] or "—",
                        "📘 Facebook": result["facebook"] or "—",
                        "📸 Instagram": result["instagram"] or "—",
                    }
                    for k, v in fields.items():
                        st.markdown(f'<div style="display:flex; gap:12px; padding:6px 0; border-bottom:1px solid rgba(99,102,241,0.08);"><span style="color:#475569; width:120px; font-size:0.82rem;">{k}</span><span style="color:#e2e8f0; font-size:0.82rem; word-break:break-all;">{v}</span></div>', unsafe_allow_html=True)

                    if result["ai_summary"]:
                        st.markdown(f'<div style="margin-top:12px; padding:10px 14px; background:rgba(99,102,241,0.08); border:1px solid rgba(99,102,241,0.2); border-radius:8px; color:#94a3b8; font-size:0.83rem;">🤖 <b style="color:#818cf8;">AI Summary:</b> {result["ai_summary"]}</div>', unsafe_allow_html=True)

                    st.markdown('</div>', unsafe_allow_html=True)

    with tab2:
        st.markdown('<p class="section-title">📥 Bulk URL Upload</p>', unsafe_allow_html=True)
        st.info("Upload a CSV with a column named **url** or **website** containing URLs to extract.")

        uploaded = st.file_uploader("Upload CSV", type=["csv"], label_visibility="collapsed")

        if uploaded:
            df_up = pd.read_csv(uploaded)
            url_col = next((c for c in df_up.columns if c.lower() in ["url", "website", "domain", "link"]), None)
            if not url_col:
                st.error("No column named 'url' or 'website' found in CSV.")
            else:
                urls = df_up[url_col].dropna().tolist()
                st.info(f"Found **{len(urls)} URLs**. Ready to extract.")

                col_opt1, col_opt2 = st.columns(2)
                with col_opt1:
                    bulk_multi = st.checkbox("Multi-page crawl", value=False, key="bulk_multi")
                with col_opt2:
                    bulk_ai = st.checkbox("AI Summaries", value=False, key="bulk_ai")

                if st.button("🚀 Start Bulk Extraction", use_container_width=True):
                    prog_bar = st.progress(0)
                    status_text = st.empty()
                    results_log = []

                    for i, url in enumerate(urls):
                        status_text.markdown(f'<div class="progress-container"><div class="progress-text">Extracting {i+1}/{len(urls)}: {url[:50]}...</div></div>', unsafe_allow_html=True)
                        prog_bar.progress((i+1) / len(urls))

                        result = extract_lead(url, multi_page=bulk_multi, max_pages=2)
                        if not result.get("error"):
                            if bulk_ai and result["company"]:
                                result["ai_summary"] = get_ai_summary(result["company"], result["website"], result["email"])
                            save_lead(result)
                            results_log.append({"url": url, "status": "✅ OK", "emails": result["email"]})
                        else:
                            results_log.append({"url": url, "status": "❌ Failed", "emails": ""})
                        time.sleep(0.5)

                    status_text.empty()
                    st.success(f"✅ Done! Processed {len(urls)} URLs.")
                    st.dataframe(pd.DataFrame(results_log), use_container_width=True)

# ══════════════════════════════════════════════
#  DASHBOARD
# ══════════════════════════════════════════════
elif page == "📊 Dashboard":
    stats = get_stats()

    st.markdown(f"""
    <div class="stat-grid">
        <div class="stat-card">
            <span class="stat-icon">🎯</span>
            <div class="stat-value">{stats['total']}</div>
            <div class="stat-label">Total Leads</div>
        </div>
        <div class="stat-card">
            <span class="stat-icon">📧</span>
            <div class="stat-value">{stats['with_email']}</div>
            <div class="stat-label">With Email</div>
            <div class="stat-delta">{'%.0f' % (stats['with_email']/max(stats['total'],1)*100)}% coverage</div>
        </div>
        <div class="stat-card">
            <span class="stat-icon">📱</span>
            <div class="stat-value">{stats['with_phone']}</div>
            <div class="stat-label">With Phone</div>
            <div class="stat-delta">{'%.0f' % (stats['with_phone']/max(stats['total'],1)*100)}% coverage</div>
        </div>
        <div class="stat-card">
            <span class="stat-icon">🔗</span>
            <div class="stat-value">{stats['with_social']}</div>
            <div class="stat-label">With Social</div>
            <div class="stat-delta">Today: {stats['today']} new</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Daily chart
    daily = get_daily_counts()
    if not daily.empty:
        st.markdown('<p class="section-title">📈 Extraction Activity</p>', unsafe_allow_html=True)
        st.bar_chart(daily.set_index("day")["count"], color="#6366f1")

    # Recent leads
    st.markdown('<p class="section-title">🕐 Recent Leads</p>', unsafe_allow_html=True)
    df = load_leads()
    if df.empty:
        st.info("No leads yet. Go to Extract to add your first one!")
    else:
        for _, row in df.head(8).iterrows():
            domain = urlparse(row.get("website","")).netloc or row.get("website","")
            initial = (row.get("company","?") or "?")[0].upper()
            date_str = str(row.get("date",""))[:10]
            email_str = row.get("email","") or "—"
            phone_str = row.get("phone","")
            social_chips = ""
            if row.get("linkedin"): social_chips += '<span class="chip">💼 LinkedIn</span> '
            if row.get("facebook"): social_chips += '<span class="chip">📘 Facebook</span> '
            if row.get("instagram"): social_chips += '<span class="chip">📸 Instagram</span> '
            phone_chip = f'<span class="chip chip-orange">📱 {phone_str[:15]}</span>' if phone_str else ""

            st.markdown(f"""
            <div class="lead-card">
                <div class="lead-avatar">{initial}</div>
                <div>
                    <div class="lead-domain">{row.get('company') or domain}</div>
                    <div class="lead-email">📧 {email_str[:60]} {phone_chip} {social_chips}</div>
                </div>
                <div class="lead-date">{date_str}</div>
            </div>
            """, unsafe_allow_html=True)

# ══════════════════════════════════════════════
#  LEADS PAGE
# ══════════════════════════════════════════════
elif page == "🗂 Leads":
    st.markdown('<p class="section-title">🗂 All Leads</p>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns([3, 1, 1])
    with col1:
        search = st.text_input("Search", placeholder="Search by company, email, URL...", label_visibility="collapsed")
    with col2:
        date_from = st.date_input("From", value=None, label_visibility="collapsed")
    with col3:
        date_to = st.date_input("To", value=None, label_visibility="collapsed")

    df = load_leads(search=search, date_from=date_from, date_to=date_to)

    if df.empty:
        st.info("No leads found.")
    else:
        st.markdown(f'<div style="color:#475569; font-size:0.82rem; margin-bottom:0.8rem;">{len(df)} leads</div>', unsafe_allow_html=True)

        # Export buttons
        col_e1, col_e2, col_e3 = st.columns([1, 1, 4])
        with col_e1:
            excel_data = to_excel(df)
            st.download_button(
                "📄 Excel",
                data=excel_data,
                file_name=f"leads_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with col_e2:
            pdf_data = to_pdf_bytes(df)
            if pdf_data:
                st.download_button(
                    "📑 PDF",
                    data=pdf_data,
                    file_name=f"leads_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )

        # Display table
        display_cols = [c for c in ["id", "company", "website", "email", "phone", "linkedin", "facebook", "instagram", "date"] if c in df.columns]
        st.dataframe(df[display_cols], use_container_width=True, hide_index=True)

        st.divider()

        # Edit / Delete
        st.markdown('<p class="section-title">✏️ Edit / 🗑 Delete Lead</p>', unsafe_allow_html=True)
        lead_ids = df["id"].tolist()
        sel_id = st.selectbox("Select Lead ID", lead_ids, format_func=lambda x: f"#{x} — {df[df['id']==x]['company'].values[0] or df[df['id']==x]['website'].values[0]}")

        if sel_id:
            sel_row = df[df["id"] == sel_id].iloc[0]
            with st.expander("✏️ Edit Lead", expanded=False):
                new_company = st.text_input("Company", value=sel_row.get("company",""))
                new_email = st.text_input("Email", value=sel_row.get("email",""))
                new_phone = st.text_input("Phone", value=sel_row.get("phone",""))
                if st.button("💾 Save Changes"):
                    update_lead(sel_id, new_company, new_email, new_phone)
                    st.success("Lead updated! Refresh to see changes.")
                    st.rerun()

            if st.button("🗑 Delete This Lead", type="primary"):
                delete_lead(sel_id)
                st.warning("Lead deleted.")
                st.rerun()

# ══════════════════════════════════════════════
#  ANALYTICS PAGE
# ══════════════════════════════════════════════
elif page == "📅 Analytics":
    st.markdown('<p class="section-title">📅 Date-wise Analytics</p>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        ana_from = st.date_input("From Date", value=None)
    with col2:
        ana_to = st.date_input("To Date", value=None)

    df_all = load_leads(date_from=ana_from, date_to=ana_to)

    if df_all.empty:
        st.info("No data in selected range.")
    else:
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total", len(df_all))
        m2.metric("With Email", int((df_all["email"] != "").sum()))
        m3.metric("With Phone", int((df_all.get("phone","") != "").sum()))
        m4.metric("With Social", int(((df_all.get("linkedin","")!="")|(df_all.get("facebook","")!="")|(df_all.get("instagram","")!="")).sum()))

        # Daily trend
        df_all["date_only"] = pd.to_datetime(df_all["date"]).dt.date
        daily = df_all.groupby("date_only").size().reset_index(name="count")
        st.markdown('<p class="section-title">📈 Daily Leads</p>', unsafe_allow_html=True)
        st.bar_chart(daily.set_index("date_only")["count"], color="#6366f1")

        # Contact coverage
        st.markdown('<p class="section-title">📊 Data Coverage</p>', unsafe_allow_html=True)
        cov_data = {
            "Email": int((df_all["email"] != "").sum()),
            "Phone": int((df_all.get("phone","") != "").sum()),
            "LinkedIn": int((df_all.get("linkedin","") != "").sum()),
            "Facebook": int((df_all.get("facebook","") != "").sum()),
            "Instagram": int((df_all.get("instagram","") != "").sum()),
            "Logo": int((df_all.get("logo_url","") != "").sum()),
        }
        cov_df = pd.DataFrame(list(cov_data.items()), columns=["Type", "Count"]).set_index("Type")
        st.bar_chart(cov_df, color="#8b5cf6")

# ══════════════════════════════════════════════
#  SETTINGS PAGE
# ══════════════════════════════════════════════
elif page == "⚙️ Settings":
    st.markdown('<p class="section-title">⚙️ Settings</p>', unsafe_allow_html=True)

    with st.expander("🗄 Database Info", expanded=True):
        stats = get_stats()
        st.markdown(f"""
        <div class="card">
            <div style="color:#64748b; font-size:0.82rem; margin-bottom:0.5rem;">Database: <span style="color:#818cf8;">leads.db</span></div>
            <div style="color:#64748b; font-size:0.82rem;">Total Records: <span style="color:#e2e8f0;">{stats['total']}</span></div>
        </div>
        """, unsafe_allow_html=True)

    with st.expander("🗑 Danger Zone", expanded=False):
        st.warning("⚠️ This will permanently delete ALL leads from the database.")
        confirm = st.text_input("Type **DELETE ALL** to confirm", placeholder="DELETE ALL")
        if st.button("🗑 Delete All Leads"):
            if confirm == "DELETE ALL":
                conn = get_conn()
                conn.execute("DELETE FROM leads")
                conn.commit()
                conn.close()
                st.success("All leads deleted.")
                st.rerun()
            else:
                st.error("Please type DELETE ALL to confirm.")

    with st.expander("📦 Export All", expanded=False):
        df_all = load_leads()
        if not df_all.empty:
            st.download_button(
                "⬇️ Download All Leads (Excel)",
                data=to_excel(df_all),
                file_name=f"all_leads_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            pdf = to_pdf_bytes(df_all)
            if pdf:
                st.download_button(
                    "⬇️ Download All Leads (PDF)",
                    data=pdf,
                    file_name=f"all_leads_{datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )

    with st.expander("ℹ️ About", expanded=False):
        st.markdown("""
        <div class="card">
            <div style="font-family:'Space Grotesk'; font-size:1.1rem; color:#818cf8; font-weight:700; margin-bottom:8px;">⚡ LeadExtractor Pro v5.0</div>
            <div style="color:#64748b; font-size:0.83rem; line-height:1.7;">
                Built by <b style="color:#e2e8f0;">Nakul Bidhuri</b> 🚀<br>
                AI-powered lead intelligence platform.<br><br>
                <b style="color:#94a3b8;">Features:</b><br>
                • Multi-page website crawling<br>
                • Email & phone extraction<br>
                • LinkedIn / Facebook / Instagram detection<br>
                • Company logo detection<br>
                • AI company summaries (Claude)<br>
                • Bulk CSV upload<br>
                • Export to Excel & PDF<br>
                • Date-wise analytics<br>
                • Edit & delete leads<br>
                • Real-time progress tracking
            </div>
        </div>
        """, unsafe_allow_html=True)
